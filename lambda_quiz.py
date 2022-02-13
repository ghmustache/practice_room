"""
ラムダ文の練習
出題（##）に対して回答を記述
openpyxlでファイル、シートを取得し、総問題数の中からrandomで問題を選択


"""
import openpyxl
import random

wb = openpyxl.load_workbook('lambda_quiz.xlsx')
ws = wb['Sheet1']

# シート内の問題数を取得
max_num = ws.max_row

# quiz開始するか質問
print('start? :')
quiz = input('quit : Enter')

# Enter key 押されるまでquiz継続
while quiz:
    rand_num = random.randint(2, max_num + 1)
    print()
    print(ws['B' + str(rand_num)].value)
    print()
    input()
    print(ws['C' + str(rand_num)].value)

    # quiz開始するか質問
    print('continue? enter any key except not enter key only')
    quiz = input()
